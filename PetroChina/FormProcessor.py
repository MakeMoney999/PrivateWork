import os,sys,time,threading,json
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import  PatternFill
from PyQt5.Qt import QFileDialog,QApplication,QMainWindow
from PyQt5.QtGui import QPixmap
import FormUI

FileDirectory='D:\\PrivateWork\\'    #文件路径
Form1Name=r'3月-6月（机关汇总表）'
Form2Name=r'Sheet1'
blackwords=['盒']
Cartridges={'CE411-413A':['CE411A','CE412A','CE413A'],'CF501-503A':['CF501A','CF502A','CF503A'],'CF401-403A':['CF401A','CF402A','CF403A']}
SplitSign_And=[',','，',' ','；',';']
SplitSign_Or=['/']
Brand_OnlyCheckName=['迪欧']
MappingName={'长尾夹':'长尾票夹','转换头':'转接头'}

class FormProcessor(FormUI.Ui_MainWindow):
    wb1 = Workbook()
    wb2 = Workbook()
    Result={}
    MatchCount=0
    UnMatch=[]
    OriginData={}
    TargetData={}
    Form1Path=''
    Form2Path=''
    Message=''
    Sheet1=''
    Sheet2=''
    mapping={}
    Switch1=False
    Switch2=False

    def getConfig(self):
        try:
            config=CFGread('config.txt')
            # config=cfgRead('config.ini')
            self.Sheet1=config[0]
            self.Sheet2=config[1]
            self.InfoShow('表名配置读取成功')
        except Exception as e:
            self.Sheet1=Form1Name
            self.Sheet2=Form2Name
            self.InfoShow('表名配置读取失败，启用默认配置')
            print(e)
        try:
            self.mapping = CovertMap(CFGread('mapping.txt'))
            self.InfoShow('替换名称配置读取成功')
            # print('mapping='+str(self.mapping))
        except:
            self.mapping = MappingName
            self.InfoShow('替换名称配置读取失败，启用默认配置')
        print(self.mapping)

    ''' 备份表 '''
    def BackupFile(self,OriginFile,TargetFile):
        cmd='copy '+OriginFile+' '+FileDirectory+TargetFile
        print(cmd)
        os.popen(cmd)
        return

    ''' 读取表文件 '''
    def GetFile(self,FilePath):
        wb = load_workbook(FilePath)
        return wb

    def GetForm(self,wb,Sheet):
        work_sheet=wb[Sheet]
        return work_sheet

    ''' 寻找表1表2的关联 '''
    def CheckForm(self,form1,form2):
        for i in range(1,9):
            self.CheckFormByLoop(form1,form2,i)
        total=form1.max_row-2
        print('Total data:',total)
        print('Totally matched Count:', self.MatchCount)
        print('Match rate:',self.MatchCount/total)
        txt='总共有'+str(total)+'条数据'
        self.InfoShow(txt)
        txt='总共匹配成功'+str(self.MatchCount)+'条数据'
        self.InfoShow(txt)
        txt='匹配率为：'+str(round(self.MatchCount/total,4)*100)+'%'
        self.InfoShow(txt)
        return self.Result

    def CheckFormByLoop(self,form1,form2,MethodID):
        for i1 in range(3, form1.max_row+1):  # 检索表1的每一行
            ''' 准备表1数据 '''
            if MatchedCheck(form1, i1) == True:
                continue
            if form1.cell(row=i1, column=3).value != None:
                self.OriginData['Brand'] = str(form1.cell(row=i1, column=3).value).upper()
            else:
                self.OriginData['Brand'] = ''
            if form1.cell(row=i1, column=4).value != None:
                self.OriginData['Model'] = str(form1.cell(row=i1, column=4).value).upper()
            else:
                self.OriginData['Model'] = ''
            self.OriginData['Name'] = str(form1.cell(row=i1, column=5).value).upper()
            for name in self.mapping:
                if name==self.OriginData['Name']:
                    print(self.mapping[name] + ' instead of ' + self.OriginData['Name'])
                    self.OriginData['MappingName']=self.mapping[name]
                    break
            else:
                self.OriginData['MappingName']='None'
            self.OriginData['Model_sub_and'] = SplitWord(self.OriginData['Model'], SplitSign_And)
            self.OriginData['Model_sub_or'] = SplitWord(self.OriginData['Model'], SplitSign_Or)

            ''' Step2 '''
            for i2 in range(3, form2.max_row):  # 检索表2的每一行
                ''' 准备表2数据 '''
                if form2.cell(row=i2, column=1).value != None:
                    form2_name = str(form2.cell(row=i2, column=1).value).upper()
                else:
                    form2_name = ''
                if form2.cell(row=i2, column=2).value != None:
                    form2_model = str(form2.cell(row=i2, column=2).value).upper()
                else:
                    form2_model = ''
                form2_name_rip = ''
                for char in range(len(form2_name)):
                    if form2_name[char] != ' ':
                        form2_name_rip += form2_name[char]

                self.TargetData = {'Name': form2_name, 'RipName': form2_name_rip, 'Model': form2_model}
                if self.CheckFormByMethod(form1, MethodID, i1, i2)==True:
                    break

    def CheckFormByMethod(self,form1,MethodID,row1,row2):
        if MethodID == 1:
        # 如果'型号'能匹配成功，则用'品牌+名称'匹配库存名称
            if (self.OriginData['Model'] in self.TargetData['Model']) and (self.OriginData['Brand'] in self.TargetData['RipName']) and (
                    self.OriginData['Name'] in self.TargetData['RipName'] or (self.OriginData['MappingName'] in self.TargetData['RipName'] and self.OriginData['MappingName']!='None')):  # 判断表1的型号与表2是否匹配
                self.MatchCount += 1
                form1.cell(row=row1, column=11).value = row2
                self.Result[self.OriginData['Name'],self.OriginData['Model'],row1] = row2
                return True
            else:
                return False

        if MethodID == 2:
            # 如果'品牌+型号+名称'与库存名称能匹配成功
            if (self.OriginData['Brand'] in self.TargetData['RipName']) and (self.OriginData['Model'] in self.TargetData['RipName']) and (
                    self.OriginData['Name'] in self.TargetData['RipName'] or (self.OriginData['MappingName'] in self.TargetData['RipName'] and self.OriginData['MappingName']!='None')):
                self.MatchCount += 1
                form1.cell(row=row1, column=11).value = row2
                self.Result[self.OriginData['Name'],self.OriginData['Model'],row1] = row2
                return True
            else:
                return False

        if MethodID == 3:
        # 如果'型号+名称'与库存名称能匹配成功
            if (self.OriginData['Model'] in self.TargetData['RipName']) and (self.OriginData['Name'] in self.TargetData['RipName'] or (self.OriginData['MappingName'] in self.TargetData['RipName'] and self.OriginData['MappingName']!='None')):
                self.MatchCount += 1
                form1.cell(row=row1, column=11).value = row2
                self.Result[self.OriginData['Name'],self.OriginData['Model'],row1] = row2
                return True
            else:
                return False

        if MethodID == 4:
        # 如果'品牌+型号'与库存名称能匹配成功
            if (self.OriginData['Brand'] in self.TargetData['RipName']) and (self.OriginData['Model'] in self.TargetData['RipName']) and self.OriginData[
            'Brand'] != '' and self.OriginData['Model'] != '':
                self.MatchCount += 1
                form1.cell(row=row1, column=11).value = row2
                self.Result[self.OriginData['Name'],self.OriginData['Model'],row1] = row2
                return True
            else:
                return False

        if MethodID == 5:
        # 拆分Model，进行组合匹配
            if ListIn(self.OriginData['Model_sub_and'], self.TargetData['RipName'], 'and', []) and (
                (self.OriginData['Name'] in self.TargetData['RipName'] or (self.OriginData['MappingName'] in self.TargetData['RipName'] and self.OriginData['MappingName']!='None')) or (self.OriginData['Brand'] in self.TargetData['RipName'])):
                self.MatchCount += 1
                form1.cell(row=row1, column=11).value = row2
                self.Result[self.OriginData['Name'],self.OriginData['Model'],row1] = row2
                return True
            else:
                return False

        if MethodID == 6:
        # 拆分Model，进行模糊匹配
            if (ListIn(self.OriginData['Model_sub_or'], self.TargetData['Model'], 'or', blackwords) or ListIn(self.OriginData['Model_sub_or'],
            self.TargetData['RipName'], 'or', blackwords)) and (self.OriginData['Name'] in self.TargetData['RipName'] or (self.OriginData['MappingName'] in self.TargetData['RipName'] and self.OriginData['MappingName']!='None')):
                self.MatchCount += 1
                form1.cell(row=row1, column=11).value = row2
                self.Result[self.OriginData['Name'],self.OriginData['Model'],row1] = row2
                return True
            else:
                return False

        if MethodID == 7:
        # 硒鼓专门匹配
            if self.OriginData['Name'] == '硒鼓':
                MatchKey = ''
                for key in Cartridges:
                    for x in self.OriginData['Model_sub_or']:
                        if x in Cartridges[key]:
                            MatchKey = key
                            break
                if (MatchKey != '') and (MatchKey in self.TargetData['Model']):
                    # print('MatchKey=',MatchKey,'form2Model=',form2_model)
                    self.MatchCount += 1
                    form1.cell(row=row1, column=11).value = row2
                    self.Result[self.OriginData['Name'],self.OriginData['Model'],row1] = row2
                    return True
                else:
                    return False

        if MethodID == 8:
        # 只匹配名称+品牌
            if (self.OriginData['Name'] in self.TargetData['RipName'] or (self.OriginData['MappingName'] in self.TargetData['RipName'] and self.OriginData['MappingName']!='None')) and (self.OriginData['Brand'] in self.TargetData['RipName']):
                self.MatchCount += 1
                form1.cell(row=row1, column=11).value = row2
                self.Result[self.OriginData['Name'],self.OriginData['Model'],row1] = row2
                return True
            else:
                return False

    def FormSave(self,wb,TargetFile):
        Path=TargetFile
        wb.save(Path)

    def CheckUnmatchedData(self):
        for i in range(3,self.Form1.max_row+1):
            if self.Form1.cell(row=i, column=11).value==None:
                self.UnMatch.append(i)
            else:
                self.Form1.cell(row=i, column=11).value = None
        txt='匹配失败'+str(len(self.UnMatch))+'条'
        self.InfoShow(txt)
        print (self.UnMatch)
        return self.UnMatch

    def CheckFormOnce(self,form1,form2):
        OriginData = {}
        count = 0
        result ={}

        ''' Step1 '''
        for i1 in range(3, form1.max_row):  # 检索表1的每一行
            ''' 准备表1数据 '''
            if form1.cell(row=i1, column=3).value!=None:
                OriginData['Brand']=str(form1.cell(row=i1, column=3).value).upper()
            else:
                OriginData['Brand']=''
            if form1.cell(row=i1, column=4).value!=None:
                OriginData['Model']=str(form1.cell(row=i1, column=4).value).upper()
            else:
                OriginData['Model']=''
            OriginData['Name']=str(form1.cell(row=i1, column=5).value).upper()
            OriginData['Model_sub_and']=SplitWord(OriginData['Model'],SplitSign_And)
            OriginData['Model_sub_or']=SplitWord(OriginData['Model'],SplitSign_Or)

            ''' Step2 '''
            for i2 in range(3,form2.max_row):   #检索表2的每一行
                ''' 准备表2数据 '''
                if form2.cell(row=i2,column=1).value!=None:
                    form2_name=str(form2.cell(row=i2,column=1).value).upper()
                else:
                    form2_name=''
                if form2.cell(row=i2,column=2).value!=None:
                    form2_model=str(form2.cell(row=i2,column=2).value).upper()
                else:
                    form2_model=''
                form2_name_rip=''
                for char in range(len(form2_name)):
                    if form2_name[char] != ' ':
                        form2_name_rip += form2_name[char]

                ''' 开始检索 '''
                # 如果'型号'能匹配成功，则用'品牌+名称'匹配库存名称
                if (OriginData['Model'] in form2_model) and (OriginData['Brand'] in form2_name) and (OriginData['Name'] in form2_name):  # 判断表1的型号与表2是否匹配
                    count += 1
                    # form1.cell(row=i1, column=11).value = i2
                    result[i1]=i2
                    break

                # 如果'品牌+型号+名称'与库存名称能匹配成功
                elif (OriginData['Brand'] in form2_name) and (OriginData['Model'] in form2_name) and (OriginData['Name'] in form2_name):
                    count += 1
                    # form1.cell(row=i1, column=12).value = i2
                    result[i1] = i2
                    break

                # 如果'型号+名称'与库存名称能匹配成功
                elif (OriginData['Model'] in form2_name) and (OriginData['Name'] in form2_name):
                    count += 1
                    # form1.cell(row=i1, column=13).value = i2
                    result[i1] = i2
                    break

                # 如果'品牌+型号'与库存名称能匹配成功
                elif (OriginData['Brand'] in form2_name) and (OriginData['Model'] in form2_name) and OriginData['Brand']!='' and OriginData['Model']!='' :
                    count += 1
                    # form1.cell(row=i1, column=14).value = i2
                    result[i1] = i2
                    break

                # 拆分Model，进行组合匹配
                elif ListIn(OriginData['Model_sub_and'],form2_name_rip,'and',[]) and ((OriginData['Name'] in form2_name_rip) or (OriginData['Brand'] in form2_name_rip)):
                    count += 1
                    # form1.cell(row=i1, column=16).value = i2
                    result[i1] = i2
                    break

                # 拆分Model，进行模糊匹配
                elif (ListIn(OriginData['Model_sub_or'],form2_model,'or',blackwords) or ListIn(OriginData['Model_sub_or'],form2_name_rip,'or',blackwords)) and (OriginData['Name'] in form2_name_rip):
                    count += 1
                    # form1.cell(row=i1, column=17).value = i2
                    result[i1] = i2
                    break

                # 型号和品牌为空时，只匹配名称
                elif OriginData['Name'] in form2_name_rip and OriginData['Model']=='' and OriginData['Brand']=='':
                    count += 1
                    # form1.cell(row=i1, column=18).value = i2
                    result[i1] = i2
                    break

                # 只匹配名称+品牌
                elif (OriginData['Name'] in form2_name_rip) and (OriginData['Brand'] in form2_name_rip):
                    count += 1
                    # form1.cell(row=i1, column=19).value = i2
                    result[i1] = i2
                    break

                # 硒鼓专门匹配
                elif OriginData['Name'] == '硒鼓':
                    MatchKey = ''
                    for key in Cartridges:
                        for x in OriginData['Model_sub_or']:
                            if x in Cartridges[key]:
                                MatchKey = key
                                break
                    if (MatchKey != '') and (MatchKey in form2_model):
                        # print('MatchKey=',MatchKey,'form2Model=',form2_model)
                        count += 1
                        # form1.cell(row=i1, column=15).value = i2
                        result[i1] = i2
                        break

        total=form1.max_row-2
        print('Total data:',total)
        print('Totally matched Count:', count)
        print('Match rate:',count/total)

        return result


    def CheckMatchResult(self,form):
        count=0
        for i in range(3, form.max_row):
            for j in range(11,20):
                if form.cell(row=i, column=j).value!=None:
                    count+=1
        print('Real matched count:',count)
        txt='实际匹配到'+str(count)+'行'
        self.InfoShow(txt)

    def CheckUnit(self,form1,form2):
        count=0
        UnitList={}
        for i in range(3, form1.max_row):
            for j in range(11, 20):
                if form1.cell(row=i, column=j).value != None:
                    form2_row=form1.cell(row=i, column=j).value
                    form1_unit=form1.cell(row=i, column=6).value
                    form2_unit=form2.cell(row=form2_row,column=3).valuess
                    if form1_unit==form2_unit:
                        form1.cell(row=i, column=21).value = 'unit matched'
                    else:
                        count+=1
                        form1.cell(row=i, column=21).value = 'unit Unmatched'
                        name1=str(form1.cell(row=i, column=5).value)+'_'+str(form1.cell(row=i, column=6).value)
                        name2=str(form2.cell(row=form2_row,column=3).value)
                        if name1 not in UnitList:
                            UnitList[name1]=name2
                        elif UnitList[name1]!=name2:
                            print("new unit of same name!!!!  " + name1 + ':'+ UnitList[name1] +'/'+name2)
        print('Unmatched count:',count)
        return UnitList

    def init(self):
        imgName='./ggzx.png'
        png = QPixmap(imgName)
        self.Logo.setPixmap(png)
        self.getConfig()
        self.Analyst_pushButton.setEnabled(False)
        self.Saveform_pushButton.setEnabled(False)
        self.SheetName1.setText(self.Sheet1)
        self.SheetName2.setText(self.Sheet2)

    def InfoShow(self,text):
        self.Message += str(text)
        self.Message += '\n'
        self.InfoShow_plainTextEdit.setPlainText(self.Message)

    def listenEvent(self):
        self.Loadform1_pushButton.clicked.connect(self.LoadForm1)
        self.Loadform2_pushButton.clicked.connect(self.LoadForm2)
        self.Saveform_pushButton.clicked.connect(self.SaveForm)
        self.Analyst_pushButton.clicked.connect(self.AnalystButtonClick)


    def LoadForm1(self):
        try:
            Form1File = QFileDialog.getOpenFileName()
            print(Form1File)
            self.Loadform1_lineEdit.setText(Form1File[0])
            self.Form1Path = Form1File[0]
            self.wb1 = self.GetFile(self.Form1Path)
            self.Form1 = self.GetForm(self.wb1 , self.Sheet1)
            self.statusbar.showMessage("采购表加载成功")
            self.InfoShow("采购表加载成功")
            self.Saveform_pushButton.setEnabled(False)
            self.Switch1 = True
        except:
            self.statusbar.showMessage("采购表加载失败，请重新加载")
            self.InfoShow("采购表加载失败，请重新加载")
            self.Switch1 = False

    def LoadForm2(self):
        try:
            Form2File = QFileDialog.getOpenFileName()
            print(Form2File)
            self.Loadform2_lineEdit.setText(Form2File[0])
            self.Form2Path = Form2File[0]
            self.wb2 = self.GetFile(self.Form2Path)
            self.Form2 = self.GetForm(self.wb2 , self.Sheet2)
            self.statusbar.showMessage("库存表加载成功")
            self.InfoShow("库存表加载成功")
            self.Switch2 = True
            if self.Loadform1_lineEdit!='':
                self.Analyst_pushButton.setEnabled(True)
                self.Saveform_pushButton.setEnabled(False)
        except:
            self.statusbar.showMessage("库存表加载失败，请重新加载")
            self.InfoShow("库存表加载失败，请重新加载")
            self.Switch2 = False

    def SaveForm(self):
        if self.Form1Path == '' or self.Form2Path == '':
            return
        try:
            NewFormDir = QFileDialog.getExistingDirectory()
            print(NewFormDir)
            name=''
            for i in range(0,len(self.Form1Path)-5):
                if self.Form1Path[i]=='/':
                    name=''
                else:
                    name+=self.Form1Path[i]
            self.newForm1=name+'_new.xlsx'
            print(self.newForm1)
            self.NewForm1Path = NewFormDir + '/' + self.newForm1
            name = ''
            for i in range(0, len(self.Form2Path) - 5):
                if self.Form2Path[i] == '/':
                    name = ''
                else:
                    name += self.Form2Path[i]
            self.newForm2 = name + '_new.xlsx'
            print(self.newForm2)
            self.NewForm2Path = NewFormDir + '/' + self.newForm2
            self.FormSave(self.wb1,self.NewForm1Path)
            self.FormSave(self.wb2,self.NewForm2Path)
            self.InfoShow(self.NewForm1Path)
            self.InfoShow(self.NewForm2Path)
            self.statusbar.showMessage("新表单保存成功")
            self.InfoShow("新表单保存成功")
        except:
            self.statusbar.showMessage("新表单保存失败，请重新保存")
            self.InfoShow("新表单保存失败，请重新保存")

    def AnalystButtonClick(self):
        if self.Switch1==False or self.Switch2==False:
            self.InfoShow('表格尚未加载成功，请检查...')
            self.statusbar.showMessage('表格尚未加载成功，请检查...')
            return
        self.Saveform_pushButton.setDisabled(True)
        self.Loadform1_pushButton.setDisabled(True)
        self.Loadform2_pushButton.setDisabled(True)
        self.Analyst_pushButton.setDisabled(True)
        try:
            self.Analyst()
            # thread = threading.Thread(target=self.Analyst)
            # thread.start()

        except:
            pass

    def Analyst(self):
        self.Result = {}
        self.MatchCount = 0
        self.OriginData = {}
        self.TargetData = {}
        self.statusbar.showMessage("分析中，请稍后...")
        self.InfoShow("分析中，请稍后...")
        try:
            result=self.CheckForm(self.Form1, self.Form2)
            print(result)
            a=rewrite().main(self.Result,self.Form1,self.Form2)
            for m in a:
                """打印需要手动写入的特殊输出结果"""
                self.statusbar.showMessage(m)
                self.InfoShow(m)
            self.CheckUnmatchedData()
            self.statusbar.showMessage("分析成功，请保存文件")
            self.InfoShow("分析成功，请保存文件")
        except Exception as e:
            print (e)
            self.statusbar.showMessage("分析失败，请联系制作者")
            self.InfoShow("分析失败")

        self.Saveform_pushButton.setEnabled(True)
        self.Loadform1_pushButton.setEnabled(True)
        self.Loadform2_pushButton.setEnabled(True)
        self.Loadform1_lineEdit.clear()
        self.Loadform2_lineEdit.clear()
        # self.Analyst_pushButton.setEnabled(True)

def SplitWord(content,symbols):
    result = []
    word = ''
    for char in range(len(content)):
        if content[char] not in symbols:
            word += content[char]
        else:
            result.append(word)
            word = ''
    if word != '':
        result.append(word)
    return result

def ListIn(origin,target,type,b_words):
    if len(origin)==0:
        return False
    elif type.lower()=='and':
        for item in origin:
            if item not in target:
                return False
        else:
            return True
    else:
        for item in origin:
            for x in b_words:
                if x in item:
                    continue
            if item in target:
                return True

def MatchedCheck(form,line):
    for x in range(11,20):
        if form.cell(row=line, column=x).value!=None:
            # print ('form row=',line,' column=',x,' matched.')
            return True
    else:
        return False

def cfgRead(file):
	with open(file,'r') as newfile:
         return json.load(newfile)

def CFGread(file):
    f=open(file,'r',encoding='utf-8')
    content=[]
    for line in f.readlines():
        if line[-1:]=='\n':
            content.append(line[:-1])
        else:
            content.append(line)
    print (content)
    f.close()
    return content

def CovertMap(map):
    result={}
    mode=1
    key=''
    value=''
    for line in map:
        mode = 1
        key=''
        value=''
        for char in line:
            if char!=',' and mode==1:
                key+=char
            elif char!=',' and mode==2:
                value+=char
            elif char==',':
                mode=2
        result[key]=value
    return result


def cfgRecord(content):
	#print (json.dumps(content))
	with open('config.ini','w') as newfile:
		newfile.write(json.dumps(content))
	return

class rewrite():
    """写表2！！将表1统计出的每个商品数量，写入表2"""
    def __init__(self):
        self.special_things_a=["小胶带","复印纸","胶带","宽胶带"]
        self.special_things_b=["纸箱","擦桌布"]  #都是5个一组

# 纸箱_组（5个）': '个', '（1组=5个）
# 擦桌布_包（5条）': '条', '（包=条）

# 三代电池_个
# 荧光笔_个': '盒', '（型号：33111，6个=1盒）
# 小胶带_卷': '筒', '（6卷=1筒）
# 复印纸_包': '箱', '（5包=1箱）
# 胶带_卷': '筒', '（6卷=1筒）
# 宽胶带_卷': '筒', '（6卷=1筒）
#双面胶_个': '袋', '（型号：30400/30411/30412，袋=卷；型号：30401,24卷=1袋;型号:30403,12卷=1袋，）

    def statistics_a(self,form1,name,x,model="0"): #计算表1中的数量关系,参数x为换算关系 ,函数有2个回调，第一个是数量，第二个是描述
        count =0
        for i in range(3, form1.max_row):#迭代表1，统计特殊商品的总数量
            if model=="0":
                if name == form1.cell(row=i,column=5).value: #没有特殊型号,直接计算累加
                    #print ("222222")
                    count+=form1.cell(row=i,column=8).value
            else:
                if name == form1.cell(row=i,column=5).value and str(model) in str(form1.cell(row=i,column=4).value):
                    #print ("3333")
                    count+=form1.cell(row=i,column=8).value
        if count%x !=0:
            txt="名称为:"+name+",总数量:"+str(count)+",型号:,"+model+",需要手填进项明细"
            #print (type(txt),txt)
            # FormProcessor().statusbar.showMessage(txt)
            # FormProcessor().InfoShow(txt)
            #print ("name=",name,"总数量",count,"型号",model,"需要手填进项明细")
            return 0,txt
        else:
            txt="名称为:"+name+",总数量:"+str(count)+",型号:,"+model+",被正常写入"
            #print ("name=",name,"总数量",count,"型号",model)
            return count/x,txt



    def get_Form1_count(self,form1,row):

        things_count=form1.cell(row=row,column=8).value
        # if form1.cell(row=row,column=5).value =="荧光笔" and form1.cell(row=row,column=4).value =="33111":
        #     things_count=0
        if form1.cell(row=row,column=5).value =="双面胶" and form1.cell(row=row,column=4).value =="30401":
            things_count=0
        elif form1.cell(row=row,column=5).value =="双面胶" and form1.cell(row=row,column=4).value =="30403":
            things_count=0
        elif form1.cell(row=row,column=5).value =="三代电池" and form1.cell(row=row,column=4).value =="5号":
            things_count=0
        elif form1.cell(row=row,column=5).value =="三代电池" and form1.cell(row=row,column=4).value =="7号":
            things_count=0    
        elif form1.cell(row=row,column=5).value in self.special_things_a:
            #print ("spec_all",things_count)
            things_count=0
            #print ("spec_all",things_count)
        elif form1.cell(row=row,column=5).value in self.special_things_b:  #b里的都是5个一组，所以统一乘5
            return things_count*5
        else:
            #print ("11111",row,things_count)
            return things_count

    def rewrite_Form2_normal(self,form2,row,things_count):  #将普通商品的数量写进表2中
        red_fill = PatternFill("solid", fgColor="FF0000")
        form2_count=form2.cell(row=row, column=10).value
        #print ("22222222,things_count")
        #print ("form2_count_a",form2_count)
        #print (type(form2_count))
        if form2_count != type(1): #如果是空就置位0
            #print("form2_count",row,type(form2_count),form2_count)
            form2_count = 0

        if things_count != type(1):
            #print("form2_count",row,type(things_count),things_count)
            things_count = 0
        
           
        #print ("22222222",things_count)    
        form2_count+= int(things_count) #出库数量做加法
        # form2.cell(row=row, column=10).value= form2_count
        # form2.cell(row=row, column=10).fill=red_fill
        form2.cell(row=row, column=10).value= form2_count
        form2.cell(row=row, column=10).fill=red_fill
        #print ("form2_count",form2_count,"form1_count",things_count)
        return form2_count

    # def rewrite_Form2_special(self,form2,row,things_count,model=0):  #需要1个查询函数，将特殊商品重写进表2中
    #     if model==0:
    #         if name==form2.cell(row=i, column=1).value:  #只找第一个
    #             form2.cell(row=i, column=10).value= things_count
    #             break
    #     else:
    #         if name==form2.cell(row=i, column=1).value and str(model) in str(form2.cell(row=i, column=2).value):
    #             form2.cell(row=i, column=10).value= things_count
    #             break
    #     return

    def rewrite_Form1(self,form1,form2,row_form1,row_form2):#将表2的名称复写进表1
        red_fill = PatternFill("solid", fgColor="FF0000")
        form1.cell(row=row_form1, column=5).value=form2.cell(row=row_form2, column=1).value  #表2的名称复写进表1
        form1.cell(row=row_form1, column=6).value=form2.cell(row=row_form2, column=3).value  #表2的单位复写进表1
        form1.cell(row=row_form1, column=5).fill=red_fill  #给复写的名称设置一个颜色
        form1.cell(row=row_form1, column=6).fill=red_fill  #给复写的单位设置一个颜色
        return 



    def main(self,matched_result,form1,form2): 
        Message_arr=[]

        #############开始全局计算特殊商品的总数量，并复写form2
        xiao_jiao_dai,string_xiao_jiao_dai=self.statistics_a(form1,"小胶带",6,"30029")  
        fu_yin_zhi,string_fu_yin_zhi=self.statistics_a(form1,"复印纸",5)
        jiao_dai,string_jiao_dai=self.statistics_a(form1,"胶带",6)
        kuai_jiao_dai,string_kuai_jiao_dai=self.statistics_a(form1,"宽胶带",6)
        shuang_mian_jiao_30401,string_shuang_mian_jiao_30401=self.statistics_a(form1,"双面胶",24,"30401")
        shuang_mian_jiao_30403,string_shuang_mian_jiao_30403=self.statistics_a(form1,"双面胶",12,"30403")
        dian_chi_5,string_dian_chi_5=self.statistics_a(form1,"三代电池",40,"5号")
        dian_chi_7,string_dian_chi_7=self.statistics_a(form1,"三代电池",40,"7号")

        Message_arr.append(string_xiao_jiao_dai)
        Message_arr.append(string_fu_yin_zhi)
        Message_arr.append(string_jiao_dai)
        Message_arr.append(string_kuai_jiao_dai)
        Message_arr.append(string_shuang_mian_jiao_30401)
        Message_arr.append(string_shuang_mian_jiao_30403)
        Message_arr.append(string_dian_chi_5)
        Message_arr.append(string_dian_chi_7)


        for form1_res,form2_row in matched_result.items(): #{('办公转椅', '', 5): 952, ('电线收纳扣', 'TLXD-A', 14): 538}
            res=self.rewrite_Form2_normal(form2,row=form2_row,things_count=self.get_Form1_count(form1,form1_res[2]))  #将表1数量，根据查询结果，写进表1
            self.rewrite_Form1(form1,form2,row_form1=form1_res[2],row_form2=form2_row)
            #复写form2，从结果中查到表2对应的行数，并复写
            if form1_res[0] == "小胶带" and form1_res[1] == "30029":
                form2.cell(row=form2_row, column=10).value= xiao_jiao_dai
            if form1_res[0] == "复印纸":
                form2.cell(row=form2_row, column=10).value= fu_yin_zhi
            if form1_res[0] == "胶带":
                form2.cell(row=form2_row, column=10).value= jiao_dai
            if form1_res[0] == "宽胶带":
                form2.cell(row=form2_row, column=10).value= kuai_jiao_dai
            if form1_res[0] == "双面胶" and form1_res[1] == "30401":
                form2.cell(row=form2_row, column=10).value= shuang_mian_jiao_30401
            if form1_res[0] == "双面胶" and form1_res[1] == "30403":
                form2.cell(row=form2_row, column=10).value= shuang_mian_jiao_30403
            if form1_res[0] == "三代电池" and form1_res[1] == "5号":
                form2.cell(row=form2_row, column=10).value= dian_chi_5
            if form1_res[0] == "三代电池" and form1_res[1] == "7号":
                form2.cell(row=form2_row, column=10).value= dian_chi_7
        return Message_arr

def RunFormProcessUI():
    app = QApplication(sys.argv)
    Window = QMainWindow()

    FP=FormProcessor()

    FP.setupUi(Window)
    FP.init()
    Window.setWindowTitle('库存表管理工具 v1.0')
    Window.show()
    FP.listenEvent()
    sys.exit(app.exec_())

if __name__=='__main__':
    RunFormProcessUI()
    # FP=FormProcessor()
    # FP.BackupFile(Form1Path, 'Form1.xlsx')
    # FP.BackupFile(Form2Path, 'Form2.xlsx')
    # Form1Path = FileDirectory + 'Form1.xlsx'
    # Form2Path = FileDirectory + 'Form2.xlsx'
    # print(Form1Path)
    # print(Form2Path)
    # time.sleep(1)
    # FP.wb1 = FP.GetFile(Form1Path)
    # FP.Form1=FP.GetForm(FP.wb1 , '3月-6月（机关汇总表）')
    # FP.wb2 = FP.GetFile(Form2Path)
    # FP.Form2=FP.GetForm(FP.wb2,'Sheet1')
    # result=FP.CheckForm(FP.Form1, FP.Form2)
    # # result=FP.CheckFormOnce(FP.Form1,FP.Form2)
    # UnmatchList=FP.CheckUnit(FP.Form1,FP.Form2)
    # FP.FormSave(FP.wb1,"NewForm1.xlsx")
    # FP.CheckMatchResult(FP.Form1)
    # #**************************
    # #写form1和fomr2
    # rewrite().main(result,FP.Form1,FP.Form2)
    # FP.FormSave(FP.wb2,"NewForm2.xlsx")
    # FP.FormSave(FP.wb1,"Result_Form2.xlsx")
    # #print('result=',result)
    # #print('UnmatchUnit=',UnmatchList)





